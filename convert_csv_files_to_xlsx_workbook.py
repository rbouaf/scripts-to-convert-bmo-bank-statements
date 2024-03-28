import os
import pandas as pd
from openpyxl import Workbook

def csv_to_excel(folder_path, excel_file):
    # Create a new Excel workbook
    wb = Workbook()

    # Loop through each CSV file in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Read the CSV file into a pandas DataFrame
            df = pd.read_csv(os.path.join(folder_path, filename))
            # Create a new sheet in the Excel workbook
            ws = wb.create_sheet(title=os.path.splitext(filename)[0])
            # Write the DataFrame to the Excel sheet
            for r_idx, row in enumerate(df.iterrows(), start=1):
                for c_idx, value in enumerate(row[1], start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the Excel workbook
    wb.save(excel_file)
    print(f"Excel file '{excel_file}' created successfully.")


# Provide the path to the folder containing CSV files
folder_path = "CSV consolidated"
# Provide the name of the Excel file to be created
excel_file = "csv_output/new_output.xlsx"

# Call the function to convert CSV files to Excel
csv_to_excel(folder_path, excel_file)